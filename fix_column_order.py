import pandas as pd
import os
import re
from openpyxl import load_workbook

def natural_sort_key(text):
    """
    Sắp xếp natural order cho columns như Q-001_1, Q-001_2, ..., Q-001_10, Q-001_11
    """
    def convert(text):
        if text.isdigit():
            return int(text)
        return text
    return [convert(c) for c in re.split('([0-9]+)', text)]

def fix_survey_file_column_order(file_path):
    """
    Fix column order in survey file from alphabetical to natural sorting
    """
    print(f"\n=== Processing {os.path.basename(file_path)} ===")
    
    try:
        # Read the Excel file with all sheets
        xls = pd.ExcelFile(file_path)
        print(f"Sheets found: {xls.sheet_names}")
        
        # Check if 'data' sheet exists
        if 'data' not in xls.sheet_names:
            print("No 'data' sheet found, skipping...")
            return False
        
        # Read the data sheet
        df_data = pd.read_excel(file_path, sheet_name='data')
        print(f"Original columns count: {len(df_data.columns)}")
        
        # Get current column order
        original_columns = df_data.columns.tolist()
        
        # Show some examples of current problematic order
        q_columns = [col for col in original_columns if 'Q-' in str(col) and '_' in str(col)]
        if q_columns:
            print("Current order examples (first 10):")
            for col in q_columns[:10]:
                print(f"  {col}")
        
        # Sort columns using natural sorting
        sorted_columns = sorted(original_columns, key=natural_sort_key)
        
        # Show examples of fixed order
        q_sorted = [col for col in sorted_columns if 'Q-' in str(col) and '_' in str(col)]
        if q_sorted:
            print("\nFixed order examples (first 10):")
            for col in q_sorted[:10]:
                print(f"  {col}")
        
        # Check if order actually changed
        if original_columns == sorted_columns:
            print("Column order is already correct, no changes needed.")
            return False
        
        # Reorder the dataframe
        df_data_sorted = df_data[sorted_columns]
        
        # Load the original workbook to preserve other sheets
        book = load_workbook(file_path)
        
        # Create backup filename
        backup_file = file_path.replace('.xlsx', '_backup.xlsx')
        print(f"Creating backup: {os.path.basename(backup_file)}")
        book.save(backup_file)
        
        # Update the data sheet with sorted columns
        if 'data' in book.sheetnames:
            del book['data']
        
        # Write the sorted data back
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_data_sorted.to_excel(writer, sheet_name='data', index=False)
        
        print(f"✅ Fixed column order in {os.path.basename(file_path)}")
        print(f"   Backup saved as: {os.path.basename(backup_file)}")
        return True
        
    except Exception as e:
        print(f"❌ Error processing {file_path}: {e}")
        return False

def main():
    survey_dir = r"D:\Downloads\FSS_TryField\data_test"
    
    if not os.path.exists(survey_dir):
        print(f"Directory not found: {survey_dir}")
        return
    
    # Find all survey Excel files
    xlsx_files = [f for f in os.listdir(survey_dir) if f.endswith('.xlsx') and not f.startswith('~')]
    
    print(f"Found {len(xlsx_files)} Excel files:")
    for file in xlsx_files:
        print(f"  - {file}")
    
    if not xlsx_files:
        print("No Excel files found to process.")
        return
    
    print(f"\n🔧 Starting column order fixes...")
    
    fixed_count = 0
    for xlsx_file in xlsx_files:
        file_path = os.path.join(survey_dir, xlsx_file)
        if fix_survey_file_column_order(file_path):
            fixed_count += 1
    
    print(f"\n✅ SUMMARY:")
    print(f"   Files processed: {len(xlsx_files)}")
    print(f"   Files fixed: {fixed_count}")
    print(f"   Files unchanged: {len(xlsx_files) - fixed_count}")
    
    if fixed_count > 0:
        print(f"\n📝 NOTE: Backup files created with '_backup.xlsx' suffix")
        print(f"   You can delete backups after verifying the fixes are correct.")

if __name__ == "__main__":
    main()